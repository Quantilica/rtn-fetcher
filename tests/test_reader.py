"""Regression tests for reader.is_metadata_row and the account hierarchy it feeds."""

from datetime import date
from types import SimpleNamespace

import openpyxl
from openpyxl.styles import Alignment
from rtn_fetcher.reader import is_metadata_row, read_sheet


def _row(*values: object) -> list[SimpleNamespace]:
    return [SimpleNamespace(value=v) for v in values]


class TestIsMetadataRow:
    def test_deflator_row_is_metadata(self):
        """Regressão: 'Deflator - IPCA base Mai/2026' é uma linha de rodapé
        das abas de valores constantes (1.1-A, 1.2-A, 1.2-B, 1.3-A, 1.4-A,
        1.5-A) com razões de deflação (~1-6), não uma conta fiscal — antes
        deste fix ela virava uma pseudo-conta de primeiro nível na árvore."""
        row = _row("Deflator - IPCA base Mai/2026", 5.54, 5.51, 5.48)
        assert is_metadata_row(row, account_columns=1)

    def test_deflator_row_is_metadata_case_insensitive_and_any_base_month(self):
        row = _row("DEFLATOR - ipca base Dez/2027", 1.02)
        assert is_metadata_row(row, account_columns=1)

    def test_existing_metadata_prefixes_still_recognized(self):
        assert is_metadata_row(_row("Obs.: Dados sujeitos à alteração."), 1)
        assert is_metadata_row(_row("Fonte: Banco Central do Brasil."), 1)
        assert is_metadata_row(_row("Memorando: PIB Nominal"), 1)
        assert is_metadata_row(_row("9/ Nota de rodapé qualquer"), 1)

    def test_real_account_row_is_not_metadata(self):
        row = _row("1.1.02    IPI", 89494.8, 89421.0, 89394.3)
        assert not is_metadata_row(row, account_columns=1)

    def test_deflator_mentioned_mid_name_is_not_metadata(self):
        # A regra é por prefixo (startswith), então um nome de conta real que
        # apenas contenha "deflator" no meio do texto não seria afetado —
        # mas ainda não existe tal conta nas planilhas RTN.
        row = _row("Alguma conta sobre deflator no meio", 1.0)
        assert not is_metadata_row(row, account_columns=1)


def _build_synthetic_1_2_b(path):
    """Minimal '1.2-B' sheet: 1 header + 2 real accounts + the deflator
    footnote row, mirroring the real STN layout closely enough to exercise
    the full read_sheet() pipeline (indentation-driven hierarchy)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1.2-B"

    ws.append(["Voltar"])
    ws.append(["Tabela 1.2-B Resultado Primário do Governo Central - Mensal"])
    ws.append(["R$ Milhões - Valores de Mai/2026 - IPCA - Acumulado em 12 meses"])
    ws.append([])
    ws.append(["Discriminação", date(2020, 1, 1), date(2020, 2, 1), date(2020, 3, 1)])
    ws.append(["1. RECEITA TOTAL", 900000.0, 910000.0, 920000.0])
    ws["A6"].alignment = Alignment(indent=0)
    ws.append(["1.1 Receita Administrada", 500000.0, 505000.0, 510000.0])
    ws["A7"].alignment = Alignment(indent=1)
    ws.append(["Deflator - IPCA base Mai/2026", 5.54, 5.51, 5.48])
    ws["A8"].alignment = Alignment(indent=0)

    wb.save(path)


def test_read_sheet_excludes_deflator_row_from_account_hierarchy(tmp_path):
    """Prova de ponta a ponta: read_sheet() não deve produzir nenhuma conta
    cujo nome comece com 'Deflator' para uma aba de valores constantes."""
    path = tmp_path / "rtn.xlsx"
    _build_synthetic_1_2_b(path)

    _data, account_hierarchy = read_sheet(path, "1.2-B")

    account_names = account_hierarchy["account_name"][1:]
    assert not any(name.lower().startswith("deflator") for name in account_names)
    # As duas contas reais continuam presentes.
    assert any("RECEITA TOTAL" in name for name in account_names)
    assert any("Receita Administrada" in name for name in account_names)
