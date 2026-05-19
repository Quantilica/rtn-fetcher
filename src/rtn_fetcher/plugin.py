# Copyright (c) 2026 Komesu, D.K.
# Licensed under the MIT License.

"""Typer plugin for quantilica-cli integration."""

from __future__ import annotations

import asyncio
import json
import logging
import sqlite3
from pathlib import Path
from typing import Annotated

import typer
from bs4 import BeautifulSoup
from rich.console import Console
from rich.logging import RichHandler
from rich.progress import (
    BarColumn,
    MofNCompleteColumn,
    Progress,
    SpinnerColumn,
    TextColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
)
from rich.rule import Rule
from rich.table import Table

from rtn_fetcher import (
    download_latest_file,
    extract_publication_metadata,
    fetch_publications_metadata,
    read_all_sheets,
)
from rtn_fetcher.cli import (
    DATA_COLUMNS,
    _bounded_download,
    get_column_names,
    get_hierarchy_columns,
    make_period_date,
    write_header,
)

app = typer.Typer(help="Resultado do Tesouro Nacional (RTN).")
export_sub = typer.Typer(help="Exportar dados RTN para diferentes formatos.")
app.add_typer(export_sub, name="export")

_DEFAULT_OUTPUT = Path("/data/rtn")
console = Console()

HIERARCHY_COLUMNS = get_hierarchy_columns(10)


def _setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.WARNING
    logging.basicConfig(
        level=level,
        format="%(message)s",
        datefmt="[%X]",
        handlers=[RichHandler(console=console, show_path=False)],
        force=True,
    )


def _make_progress() -> Progress:
    return Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        TimeRemainingColumn(),
        console=console,
    )


def _resolve_publications(
    output: Path, metadata: Path | None, force: bool
) -> list[dict]:
    """Resolve a lista de publicações RTN (de arquivo ou via rede)."""
    if metadata is not None:
        if not metadata.exists():
            console.print(
                f"[red]Erro:[/red] metadados não encontrados: {metadata}"
            )
            raise typer.Exit(code=1)
        return json.loads(metadata.read_text(encoding="utf-8"))

    out_html = output / "metadata.html"
    if out_html.exists() and not force:
        metadata_html = out_html.read_text(encoding="utf-8")
    else:
        with console.status("[cyan]Buscando metadados RTN...[/cyan]"):
            metadata_html = fetch_publications_metadata()
        out_html.write_text(metadata_html, encoding="utf-8")
    soup = BeautifulSoup(metadata_html, "html.parser")
    return extract_publication_metadata(soup)


def _sync_publications(
    output: Path,
    publications: list[dict],
    concurrency: int,
) -> tuple[int, int, int]:
    """Baixar todas as publicações RTN. Retorna (ok, falhas, pulados)."""
    tasks_info = [
        (pub, filename, url)
        for pub in publications
        for filename, url in pub.get("links", {}).items()
    ]
    ok = failed = skipped = 0

    with _make_progress() as progress:
        task = progress.add_task(
            "[cyan]Baixando arquivos RTN...[/cyan]", total=len(tasks_info)
        )

        def on_done(filename: str, result: str) -> None:
            nonlocal ok, failed, skipped
            if result == "ok":
                ok += 1
            elif result == "skipped":
                skipped += 1
            else:
                failed += 1
            progress.update(
                task,
                advance=1,
                description=(
                    f"[green]{ok}✓[/green]  "
                    f"[red]{failed}✗[/red]  "
                    f"[dim]{skipped} skip[/dim]"
                ),
            )

        async def _run() -> None:
            from quantilica_core.http import BROWSER_HEADERS, AsyncHttpClient

            semaphore = asyncio.Semaphore(concurrency)
            client = AsyncHttpClient(timeout=600.0, headers=BROWSER_HEADERS)
            await asyncio.gather(
                *[
                    _bounded_download(
                        pub,
                        fn,
                        url,
                        client,
                        output,
                        "utf-8",
                        semaphore,
                        on_done=on_done,
                    )
                    for pub, fn, url in tasks_info
                ]
            )

        asyncio.run(_run())

    return ok, failed, skipped


@app.command("sync")
def cmd_sync(
    output: Annotated[
        Path, typer.Option("-o", "--output", help="Diretório de destino")
    ] = _DEFAULT_OUTPUT,
    latest: Annotated[
        bool,
        typer.Option(
            "--latest", help="Baixar apenas a série histórica mais recente"
        ),
    ] = False,
    metadata: Annotated[
        Path | None,
        typer.Option(
            "--metadata",
            help="JSON de metadados existente (dispensa busca online)",
        ),
    ] = None,
    concurrency: Annotated[
        int, typer.Option("--concurrency", help="Downloads simultâneos")
    ] = 4,
    force: Annotated[
        bool,
        typer.Option(
            "--force", help="Rebaixar metadados mesmo se já existirem"
        ),
    ] = False,
    dry_run: Annotated[
        bool, typer.Option("--dry-run", help="Listar arquivos sem baixar")
    ] = False,
    verbose: Annotated[
        bool, typer.Option("--verbose", help="Logs detalhados")
    ] = False,
) -> None:
    """Sincronizar publicações RTN com o diretório local."""
    _setup_logging(verbose)
    output.mkdir(parents=True, exist_ok=True)

    if latest:
        with console.status(
            "[cyan]Baixando arquivo RTN mais recente...[/cyan]"
        ):
            filepath = download_latest_file(output)
        console.print(
            f"[green]✓[/green] Arquivo RTN mais recente:"
            f" [bold]{filepath}[/bold]"
        )
        return

    publications = _resolve_publications(output, metadata, force)
    tasks_info = [
        (pub, filename, url)
        for pub in publications
        for filename, url in pub.get("links", {}).items()
    ]

    if dry_run:
        t = Table(show_header=True, header_style="bold")
        t.add_column("Arquivo", style="cyan")
        t.add_column("Período", style="dim")
        for pub, fn, _ in tasks_info:
            t.add_row(
                fn,
                f"{pub.get('ano_publicacao')}"
                f"-{pub.get('mes_publicacao'):0>2}",
            )
        console.print(t)
        console.print(f"[bold]Total:[/bold] {len(tasks_info)} arquivo(s)")
        return

    try:
        ok, failed, skipped = _sync_publications(
            output, publications, concurrency
        )
    except KeyboardInterrupt:
        console.print("[yellow]Download cancelado.[/yellow]")
        raise typer.Exit(code=130) from None

    if failed:
        console.print(
            f"[yellow]⚠[/yellow]  {ok} OK · "
            f"[red]{failed} falha(s)[/red] · {skipped} pulados"
        )
    else:
        console.print(
            f"[green]✓[/green]  [bold]{ok}[/bold] baixados · "
            f"[dim]{skipped} pulados[/dim]"
        )


def _export_excel(filepath: Path, save_as: Path) -> None:
    """Exportar um arquivo RTN para Excel."""
    from openpyxl import Workbook

    with console.status(
        f"[cyan]Lendo planilhas de {filepath.name}...[/cyan]"
    ):
        results = read_all_sheets(filepath)

    wb = Workbook()
    wb.remove(wb.active)
    ws_data = wb.create_sheet(title="rtn_data")
    write_header(ws_data, DATA_COLUMNS)
    ws_acc = wb.create_sheet(title="rtn_accounts")
    write_header(ws_acc, HIERARCHY_COLUMNS)

    data_row = acc_row = 2
    for sheet_name, (data_tbl, accounts_tbl, sheet_title) in results.items():
        data_header = get_column_names(data_tbl)
        for row in data_tbl.iter_rows():
            if row == data_header:
                continue
            rd = dict(zip(data_header, row, strict=False))
            period_date = make_period_date(
                rd.get("year"), rd.get("month"), rd.get("quarter")
            )
            for col_idx, value in enumerate(
                [
                    f"{sheet_name}|{rd.get('account')}",
                    period_date,
                    rd.get("value"),
                ],
                1,
            ):
                ws_data.cell(row=data_row, column=col_idx, value=value)
            data_row += 1

        acc_header = get_column_names(accounts_tbl)
        for row in accounts_tbl.iter_rows():
            if row == acc_header:
                continue
            rd = dict(zip(acc_header, row, strict=False))
            account_code = rd.get("account_code")
            p_values = [rd.get(f"P_{i}") for i in range(1, 11)]
            for col_idx, value in enumerate(
                [
                    f"{sheet_name}|{account_code}",
                    sheet_name,
                    sheet_title,
                    account_code,
                    rd.get("account_name"),
                    rd.get("account_level"),
                    *p_values,
                ],
                1,
            ):
                ws_acc.cell(row=acc_row, column=col_idx, value=value)
            acc_row += 1

    wb.save(save_as)
    console.print(f"[green]✓[/green] Salvo em [bold]{save_as}[/bold]")


def _export_sqlite(filepath: Path, save_as: Path, force: bool) -> None:
    """Exportar um arquivo RTN para SQLite."""
    with console.status(
        f"[cyan]Lendo planilhas de {filepath.name}...[/cyan]"
    ):
        results = read_all_sheets(filepath)

    if save_as.exists():
        if not force:
            typer.confirm(
                f"O banco '{save_as}' já existe. Sobrescrever?",
                abort=True,
            )
        save_as.unlink()

    conn = sqlite3.connect(save_as)
    cursor = conn.cursor()
    hierarchy_cols = [f"P_{i}" for i in range(1, 11)]
    cols_sql = ", ".join(f"{col} TEXT" for col in hierarchy_cols)
    cursor.execute("CREATE TABLE rtn_data (key TEXT, date TEXT, value REAL)")
    cursor.execute(
        "CREATE TABLE rtn_accounts (key TEXT, sheet TEXT, sheet_title TEXT, "
        "account_code TEXT, account_name TEXT, account_level INTEGER, "
        f"{cols_sql})"
    )

    for sheet_name, (data_tbl, accounts_tbl, sheet_title) in results.items():
        data_header = get_column_names(data_tbl)
        for row in data_tbl.iter_rows():
            if row == data_header:
                continue
            rd = dict(zip(data_header, row, strict=False))
            period_date = make_period_date(
                rd.get("year"), rd.get("month"), rd.get("quarter")
            )
            cursor.execute(
                "INSERT INTO rtn_data VALUES (?, ?, ?)",
                (
                    f"{sheet_name}|{rd.get('account')}",
                    period_date,
                    rd.get("value"),
                ),
            )

        acc_header = get_column_names(accounts_tbl)
        for row in accounts_tbl.iter_rows():
            if row == acc_header:
                continue
            rd = dict(zip(acc_header, row, strict=False))
            account_code = rd.get("account_code")
            p_values = [rd.get(f"P_{i}") for i in range(1, 11)]
            placeholders = ", ".join(["?"] * 10)
            cursor.execute(
                "INSERT INTO rtn_accounts (key, sheet, sheet_title, "
                "account_code, account_name, account_level, "
                f"{', '.join(hierarchy_cols)}) "
                f"VALUES (?, ?, ?, ?, ?, ?, {placeholders})",
                [
                    f"{sheet_name}|{account_code}",
                    sheet_name,
                    sheet_title,
                    account_code,
                    rd.get("account_name"),
                    rd.get("account_level"),
                    *p_values,
                ],
            )

    conn.commit()
    conn.close()
    console.print(
        f"[green]✓[/green] Banco de dados salvo em [bold]{save_as}[/bold]"
    )


def _resolve_rtn_file(output: Path, file: Path | None) -> Path:
    """Devolver o arquivo RTN local, baixando o mais recente se preciso."""
    if file is not None:
        return file
    with console.status(
        "[cyan]Baixando arquivo RTN mais recente...[/cyan]"
    ):
        return download_latest_file(output)


@export_sub.command("excel")
def cmd_export_excel(
    output: Annotated[
        Path,
        typer.Option("-o", "--output", help="Diretório com arquivos RTN"),
    ] = _DEFAULT_OUTPUT,
    save_as: Annotated[
        Path, typer.Option("--save-as", help="Arquivo Excel de saída")
    ] = Path("rtn_processed.xlsx"),
    file: Annotated[
        Path | None,
        typer.Option(
            "--file",
            help="Arquivo RTN local (dispensa download automático)",
        ),
    ] = None,
    verbose: Annotated[
        bool, typer.Option("--verbose", help="Logs detalhados")
    ] = False,
) -> None:
    """Exportar dados RTN para Excel."""
    _setup_logging(verbose)
    output.mkdir(parents=True, exist_ok=True)
    _export_excel(_resolve_rtn_file(output, file), save_as)


@export_sub.command("sqlite")
def cmd_export_sqlite(
    output: Annotated[
        Path,
        typer.Option("-o", "--output", help="Diretório com arquivos RTN"),
    ] = _DEFAULT_OUTPUT,
    save_as: Annotated[
        Path, typer.Option("--save-as", help="Arquivo SQLite de saída")
    ] = Path("rtn_data.db"),
    file: Annotated[
        Path | None,
        typer.Option(
            "--file",
            help="Arquivo RTN local (dispensa download automático)",
        ),
    ] = None,
    force: Annotated[
        bool,
        typer.Option(
            "--force",
            "-f",
            help="Sobrescrever banco existente sem confirmação",
        ),
    ] = False,
    verbose: Annotated[
        bool, typer.Option("--verbose", help="Logs detalhados")
    ] = False,
) -> None:
    """Exportar dados RTN para SQLite."""
    _setup_logging(verbose)
    output.mkdir(parents=True, exist_ok=True)
    _export_sqlite(_resolve_rtn_file(output, file), save_as, force)


@app.command("pipeline")
def cmd_pipeline(
    output: Annotated[
        Path, typer.Option("-o", "--output", help="Diretório de destino")
    ] = _DEFAULT_OUTPUT,
    fmt: Annotated[
        str,
        typer.Option("--format", help="Formato de exportação (excel|sqlite)"),
    ] = "excel",
    save_as: Annotated[
        Path | None,
        typer.Option("--save-as", help="Arquivo de saída da exportação"),
    ] = None,
    concurrency: Annotated[
        int, typer.Option("--concurrency", help="Downloads simultâneos")
    ] = 4,
    force: Annotated[
        bool,
        typer.Option("--force", help="Sobrescrever saídas existentes"),
    ] = False,
    verbose: Annotated[
        bool, typer.Option("--verbose", help="Logs detalhados")
    ] = False,
) -> None:
    """Pipeline completo RTN (sync → export)."""
    _setup_logging(verbose)
    if fmt not in ("excel", "sqlite"):
        console.print(
            f"[red]Erro:[/red] formato inválido '{fmt}' (use excel ou sqlite)"
        )
        raise typer.Exit(code=1)
    output.mkdir(parents=True, exist_ok=True)

    console.print(Rule("[bold]Passo 1/2: Sincronização[/bold]"))
    publications = _resolve_publications(output, None, force)
    ok, failed, skipped = _sync_publications(output, publications, concurrency)
    console.print(
        f"[green]✓[/green] {ok} baixados · {failed} falhas · {skipped} skip"
    )

    console.print(Rule("[bold]Passo 2/2: Exportação[/bold]"))
    filepath = _resolve_rtn_file(output, None)
    if fmt == "excel":
        _export_excel(filepath, save_as or Path("rtn_processed.xlsx"))
    else:
        _export_sqlite(filepath, save_as or Path("rtn_data.db"), force)
