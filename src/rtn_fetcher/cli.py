#!/usr/bin/env python3
"""Unified CLI tool for RTN data fetching and export operations.

Usage examples:
  rtn-fetcher download
  rtn-fetcher latest
  rtn-fetcher export excel
  rtn-fetcher export sqlite
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sqlite3
import sys
from collections.abc import Callable
from datetime import date
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from quantilica_core.http import BROWSER_HEADERS, AsyncHttpClient
from quantilica_core.logging import configure_cli_logging

from rtn_fetcher import (
    Tbl,
    __version__,
    download_latest_file,
    extract_publication_metadata,
    fetch_publications_metadata,
    logger,
    read_all_sheets,
)
from rtn_fetcher.fetcher import download_publication_link

DEFAULT_DATA_DIR = Path("/data/rtn")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")

DATA_COLUMNS = ["key", "date", "value"]


def get_hierarchy_columns(p_column_count: int) -> list[str]:
    """Build the hierarchy column headers based on the number of P_ columns."""
    return [
        "key",
        "sheet",
        "sheet_title",
        "account_code",
        "account_name",
        "account_level",
        *(f"P_{i}" for i in range(1, p_column_count + 1)),
    ]


def make_period_date(year, month, quarter) -> date | None:
    if year is None:
        return None
    if quarter is not None:
        return date(int(year), (int(quarter) - 1) * 3 + 1, 1)
    if month is not None:
        return date(int(year), int(month), 1)
    return date(int(year), 1, 1)


def write_header(ws, columns: list[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def get_column_names(table: Tbl) -> list[str]:
    return [col[0] for col in table.data]


def get_p_column_count(columns: list[str]) -> int:
    """Count how many P_i hierarchy columns exist in the column list."""
    count = 0
    for i in range(1, 11):
        if f"P_{i}" in columns:
            count += 1
        else:
            break
    return count


async def _bounded_download(
    pub: dict,
    filename: str,
    url: str,
    client: AsyncHttpClient,
    dest_root: Path,
    encoding: str,
    semaphore: asyncio.Semaphore,
    *,
    on_done: Callable[[str, str], None] | None = None,
) -> None:
    """Download one publication link, gated by ``semaphore``."""
    ano = pub.get("ano_publicacao")
    mes = pub.get("mes_publicacao")
    dest_dir = dest_root / f"{ano}-{mes:0>2}"
    dest_dir.mkdir(parents=True, exist_ok=True)

    dest_file = dest_dir / filename
    if dest_file.exists():
        logger.debug(f"Skipped {filename} (already exists)")
        if on_done is not None:
            on_done(filename, "skipped")
        return

    async with semaphore:
        logger.debug(f"Downloading {filename}...")
        try:
            await download_publication_link(
                client, url, dest_file, text_encoding=encoding
            )
            if on_done is not None:
                on_done(filename, "ok")
        except Exception as exc:
            logger.error(f"Failed to download {url}: {exc}")
            if on_done is not None:
                on_done(filename, "failed")


def _load_publications(
    dest: Path, metadata_path: Path | None, force: bool
) -> list[dict]:
    """Fetch or load RTN publication metadata, caching HTML to disk."""
    if metadata_path is not None:
        if not metadata_path.exists():
            print(
                f"Erro: metadados não encontrados: {metadata_path}",
                file=sys.stderr,
            )
            sys.exit(1)
        with metadata_path.open("r", encoding="utf-8") as f:
            return json.load(f)

    out_html = dest / "metadata.html"
    if out_html.exists() and not force:
        logger.info("Using cached metadata.")
        with out_html.open("r", encoding="utf-8") as f:
            metadata_html = f.read()
    else:
        logger.info("Fetching RTN metadata...")
        metadata_html = fetch_publications_metadata()
        out_html.parent.mkdir(parents=True, exist_ok=True)
        with out_html.open("w", encoding="utf-8") as f:
            f.write(metadata_html)

    soup = BeautifulSoup(metadata_html, "html.parser")
    publications = extract_publication_metadata(soup)
    logger.info(f"Found {len(publications)} publications.")
    return publications


def cmd_download(args: argparse.Namespace) -> None:
    """Fetch metadata and download all RTN publication files."""
    dest = args.output
    dest.mkdir(parents=True, exist_ok=True)

    if getattr(args, "latest", False):
        filepath = download_latest_file(dest)
        logger.info(f"Latest RTN file: {filepath}")
        return

    publications = _load_publications(dest, args.metadata, args.force)
    total = sum(len(pub.get("links", {})) for pub in publications)
    ok = failed = skipped = 0

    def on_done(filename: str, result: str) -> None:
        nonlocal ok, failed, skipped
        if result == "ok":
            ok += 1
        elif result == "skipped":
            skipped += 1
        else:
            failed += 1
        count = ok + failed + skipped
        logger.info("[%d/%d] %s (%s)", count, total, filename, result)

    async def _run_async_downloads() -> None:
        semaphore = asyncio.Semaphore(args.concurrency)
        client = AsyncHttpClient(timeout=600.0, headers=BROWSER_HEADERS)
        tasks = [
            _bounded_download(
                pub,
                filename,
                url,
                client,
                dest,
                "utf-8",
                semaphore,
                on_done=on_done,
            )
            for pub in publications
            for filename, url in pub.get("links", {}).items()
        ]
        if tasks:
            await asyncio.gather(*tasks)

    asyncio.run(_run_async_downloads())


def cmd_pipeline(args: argparse.Namespace) -> None:
    """Run the full RTN pipeline: sync then export."""
    if args.format not in ("excel", "sqlite"):
        print(
            f"Erro: formato inválido '{args.format}' (use excel ou sqlite)",
            file=sys.stderr,
        )
        sys.exit(1)
    logger.info("=== Passo 1/2: sincronização ===")
    args.latest = False
    cmd_download(args)
    logger.info("=== Passo 2/2: exportação ===")
    if args.format == "excel":
        if args.save_as is None:
            args.save_as = Path("rtn_processed.xlsx")
        cmd_export_excel(args)
    else:
        if args.save_as is None:
            args.save_as = Path("rtn_data.db")
        cmd_export_sqlite(args)


def cmd_export_excel(args: argparse.Namespace) -> None:
    """Export RTN data to Excel."""
    data_dir = args.output
    data_dir.mkdir(parents=True, exist_ok=True)

    logger.info("Checking for latest RTN file...")
    filepath = download_latest_file(data_dir)
    logger.info(f"Using file: {filepath}")

    logger.info(f"Reading all sheets from {filepath}...")
    results = read_all_sheets(filepath)

    # Detect max hierarchy depth across all sheets
    max_p_columns = 0
    for _, (_, accounts_tbl, _) in results.items():
        acc_header = get_column_names(accounts_tbl)
        p_count = get_p_column_count(acc_header)
        max_p_columns = max(max_p_columns, p_count)

    output_path = Path(args.save_as)
    wb = Workbook()
    wb.remove(wb.active)

    ws_data = wb.create_sheet(title="rtn_data")
    write_header(ws_data, DATA_COLUMNS)
    data_row = 2

    ws_acc = wb.create_sheet(title="rtn_accounts")
    hierarchy_columns = get_hierarchy_columns(max_p_columns)
    write_header(ws_acc, hierarchy_columns)
    acc_row = 2

    for sheet_name, (data_tbl, accounts_tbl, sheet_title) in results.items():
        logger.info(f"Processing sheet {sheet_name}...")

        data_header = get_column_names(data_tbl)
        for row in data_tbl.iter_rows():
            if row == data_header:
                continue
            row_dict = dict(zip(data_header, row, strict=False))
            account = row_dict.get("account")
            period_date = make_period_date(
                row_dict.get("year"),
                row_dict.get("month"),
                row_dict.get("quarter"),
            )
            data_values = [
                f"{sheet_name}|{account}",
                period_date,
                row_dict.get("value"),
            ]
            for col_idx, value in enumerate(data_values, start=1):
                ws_data.cell(row=data_row, column=col_idx, value=value)
            data_row += 1

        acc_header = get_column_names(accounts_tbl)
        p_column_count = get_p_column_count(acc_header)
        for row in accounts_tbl.iter_rows():
            if row == acc_header:
                continue
            row_dict = dict(zip(acc_header, row, strict=False))
            account_code = row_dict.get("account_code")
            p_values = [row_dict.get(f"P_{i}") for i in range(1, p_column_count + 1)]
            for col_idx, value in enumerate(
                [
                    f"{sheet_name}|{account_code}",
                    sheet_name,
                    sheet_title,
                    account_code,
                    row_dict.get("account_name"),
                    row_dict.get("account_level"),
                    *p_values,
                ],
                start=1,
            ):
                ws_acc.cell(row=acc_row, column=col_idx, value=value)
            acc_row += 1

    wb.save(output_path)
    logger.info(f"Saved to {output_path}")


def cmd_export_sqlite(args: argparse.Namespace) -> None:
    """Export RTN data to SQLite database."""
    data_dir = args.output
    data_dir.mkdir(parents=True, exist_ok=True)

    logger.info("Checking for latest RTN file...")
    filepath = download_latest_file(data_dir)
    logger.info(f"Using file: {filepath}")

    logger.info(f"Reading all sheets from {filepath}...")
    results = read_all_sheets(filepath)

    db_path = Path(args.save_as)
    if db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute(
        """
        CREATE TABLE rtn_data (
            key   TEXT,
            date  TEXT,
            value REAL
        )
    """
    )

    # Detect max hierarchy depth across all sheets
    max_p_columns = 0
    for _, (_, accounts_tbl, _) in results.items():
        acc_header = get_column_names(accounts_tbl)
        p_count = get_p_column_count(acc_header)
        max_p_columns = max(max_p_columns, p_count)

    hierarchy_cols = [f"P_{i}" for i in range(1, max_p_columns + 1)]
    cols_sql = ", ".join(f"{col} TEXT" for col in hierarchy_cols)
    cursor.execute(
        f"""
        CREATE TABLE rtn_accounts (
            key           TEXT,
            sheet         TEXT,
            sheet_title   TEXT,
            account_code  TEXT,
            account_name  TEXT,
            account_level INTEGER,
            {cols_sql}
        )
    """
    )

    for sheet_name, (data_tbl, accounts_tbl, sheet_title) in results.items():
        logger.info(f"Processing sheet {sheet_name}...")

        data_header = get_column_names(data_tbl)
        for row in data_tbl.iter_rows():
            if row == data_header:
                continue
            row_dict = dict(zip(data_header, row, strict=False))
            account = row_dict.get("account")
            period_date = make_period_date(
                row_dict.get("year"),
                row_dict.get("month"),
                row_dict.get("quarter"),
            )
            cursor.execute(
                "INSERT INTO rtn_data (key, date, value) VALUES (?, ?, ?)",
                (
                    f"{sheet_name}|{account}",
                    period_date,
                    row_dict.get("value"),
                ),
            )

        acc_header = get_column_names(accounts_tbl)
        p_column_count = get_p_column_count(acc_header)
        for row in accounts_tbl.iter_rows():
            if row == acc_header:
                continue
            row_dict = dict(zip(acc_header, row, strict=False))
            account_code = row_dict.get("account_code")
            p_values = [row_dict.get(f"P_{i}") for i in range(1, p_column_count + 1)]
            # Pad with None to match the max_p_columns for consistency
            p_values.extend([None] * (max_p_columns - p_column_count))
            placeholders = ", ".join(["?"] * max_p_columns)
            cursor.execute(
                f"INSERT INTO rtn_accounts (key, sheet, sheet_title, "
                f"account_code, account_name, account_level, "
                f"{', '.join(hierarchy_cols)}) "
                f"VALUES (?, ?, ?, ?, ?, ?, {placeholders})",
                [
                    f"{sheet_name}|{account_code}",
                    sheet_name,
                    sheet_title,
                    account_code,
                    row_dict.get("account_name"),
                    row_dict.get("account_level"),
                    *p_values,
                ],
            )

    conn.commit()
    conn.close()
    logger.info(f"Database saved to {db_path}")


def get_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="rtn-fetcher",
        description="RTN data helper CLI - fetch and export utilities",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {__version__}",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        default=False,
        help="Exibir logs detalhados (DEBUG)",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_dl = sub.add_parser(
        "sync",
        help="Sincronizar publicações RTN com o diretório local",
    )
    p_dl.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_DATA_DIR,
        help=f"Diretório de destino (padrão: {DEFAULT_DATA_DIR})",
    )
    p_dl.add_argument(
        "--metadata",
        type=Path,
        default=None,
        metavar="FILE",
        help="JSON de metadados existente (dispensa busca online)",
    )
    p_dl.add_argument(
        "--force",
        action="store_true",
        help="Rebaixar metadados mesmo se já existirem",
    )
    p_dl.add_argument(
        "--concurrency",
        type=int,
        default=4,
        help="Downloads simultâneos (padrão: 4)",
    )
    p_dl.add_argument(
        "--latest",
        action="store_true",
        help="Baixar apenas a série histórica RTN mais recente",
    )
    p_dl.set_defaults(func=cmd_download)

    p_pipeline = sub.add_parser(
        "pipeline",
        help="Pipeline completo RTN (sync -> export)",
    )
    p_pipeline.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_DATA_DIR,
        help=f"Diretório de destino (padrão: {DEFAULT_DATA_DIR})",
    )
    p_pipeline.add_argument(
        "--format",
        choices=["excel", "sqlite"],
        default="excel",
        help="Formato de exportação (padrão: excel)",
    )
    p_pipeline.add_argument(
        "--save-as",
        dest="save_as",
        type=Path,
        default=None,
        help="Arquivo de saída da exportação",
    )
    p_pipeline.add_argument(
        "--metadata",
        type=Path,
        default=None,
        metavar="FILE",
        help="JSON de metadados existente (dispensa busca online)",
    )
    p_pipeline.add_argument(
        "--force",
        action="store_true",
        help="Rebaixar metadados mesmo se já existirem",
    )
    p_pipeline.add_argument(
        "--concurrency",
        type=int,
        default=4,
        help="Downloads simultâneos (padrão: 4)",
    )
    p_pipeline.set_defaults(func=cmd_pipeline)

    p_export = sub.add_parser(
        "export",
        help="Exportar dados RTN para diferentes formatos",
    )
    export_sub = p_export.add_subparsers(dest="export_command", required=True)

    p_excel = export_sub.add_parser("excel", help="Exportar para Excel")
    p_excel.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_DATA_DIR,
        help=f"Diretório com arquivos RTN (padrão: {DEFAULT_DATA_DIR})",
    )
    p_excel.add_argument(
        "--save-as",
        dest="save_as",
        type=Path,
        default=Path("rtn_processed.xlsx"),
        help="Arquivo Excel de saída (padrão: rtn_processed.xlsx)",
    )
    p_excel.set_defaults(func=cmd_export_excel)

    p_sqlite = export_sub.add_parser("sqlite", help="Exportar para SQLite")
    p_sqlite.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_DATA_DIR,
        help=f"Diretório com arquivos RTN (padrão: {DEFAULT_DATA_DIR})",
    )
    p_sqlite.add_argument(
        "--save-as",
        dest="save_as",
        type=Path,
        default=Path("rtn_data.db"),
        help="Arquivo SQLite de saída (padrão: rtn_data.db)",
    )
    p_sqlite.set_defaults(func=cmd_export_sqlite)

    return parser


def main(argv: list[str] | None = None) -> None:
    parser = get_parser()
    args = parser.parse_args(argv)
    configure_cli_logging(verbose=args.verbose)
    args.func(args)


if __name__ == "__main__":
    main()
