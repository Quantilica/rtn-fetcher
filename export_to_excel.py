"""Script to download the latest RTN data and export it to an Excel file."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from rtnpy import Tbl, download_latest_file, read_all_sheets

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")

DATA_COLUMNS = ["key", "date", "value"]
HIERARCHY_COLUMNS = (
    ["key", "sheet", "sheet_title", "account_code", "account_name", "account_level"]
    + [f"P_{i}" for i in range(1, 11)]
)


def make_period_date(year, month, quarter) -> date | None:
    if year is None:
        return None
    if quarter is not None:
        return date(int(year), (int(quarter) - 1) * 3 + 1, 1)
    if month is not None:
        return date(int(year), int(month), 1)
    return date(int(year), 1, 1)


def write_header(ws, columns: list[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def get_column_names(table: Tbl) -> list[str]:
    return [col[0] for col in table.data]


def save_to_excel(
    results: dict[str, tuple[Tbl, Tbl, str | None]],
    output_path: Path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_data = wb.create_sheet(title="rtn_data")
    write_header(ws_data, DATA_COLUMNS)
    data_row = 2

    ws_acc = wb.create_sheet(title="rtn_accounts")
    write_header(ws_acc, HIERARCHY_COLUMNS)
    acc_row = 2

    for sheet_name, (data_tbl, accounts_tbl, sheet_title) in results.items():
        print(f"Processing sheet {sheet_name}...")

        data_header = get_column_names(data_tbl)
        for row in data_tbl.iter_rows():
            if row == data_header:
                continue
            row_dict = dict(zip(data_header, row))
            account = row_dict.get("account")
            period_date = make_period_date(
                row_dict.get("year"),
                row_dict.get("month"),
                row_dict.get("quarter"),
            )
            for col_idx, value in enumerate(
                [f"{sheet_name}|{account}", period_date, row_dict.get("value")],
                start=1,
            ):
                ws_data.cell(row=data_row, column=col_idx, value=value)
            data_row += 1

        acc_header = get_column_names(accounts_tbl)
        for row in accounts_tbl.iter_rows():
            if row == acc_header:
                continue
            row_dict = dict(zip(acc_header, row))
            account_code = row_dict.get("account_code")
            p_values = [row_dict.get(f"P_{i}") for i in range(1, 11)]
            for col_idx, value in enumerate(
                [
                    f"{sheet_name}|{account_code}",
                    sheet_name,
                    sheet_title,
                    account_code,
                    row_dict.get("account_name"),
                    row_dict.get("account_level"),
                    *p_values,
                ],
                start=1,
            ):
                ws_acc.cell(row=acc_row, column=col_idx, value=value)
            acc_row += 1

    wb.save(output_path)
    print(f"Saved to {output_path}")


def main() -> None:
    data_dir = Path("data")
    data_dir.mkdir(exist_ok=True)

    print("Checking for latest RTN file...")
    filepath = download_latest_file(data_dir)

    if filepath is None:
        xlsx_files = sorted(data_dir.glob("rtn_*.xlsx"), reverse=True)
        if not xlsx_files:
            print("No RTN files found and could not download latest.")
            return
        filepath = xlsx_files[0]
        print(f"Using existing file: {filepath}")
    else:
        print(f"Downloaded: {filepath}")

    print(f"Reading all sheets from {filepath}...")
    results = read_all_sheets(filepath)

    output_path = Path("rtn_processed.xlsx")
    save_to_excel(results, output_path)


if __name__ == "__main__":
    main()
