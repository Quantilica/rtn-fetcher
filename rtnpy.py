#!/usr/bin/env python3
"""Unified CLI tool for RTN data fetching and export operations.

Provides these subcommand groups:
- fetch: metadata, download, latest
- export: excel, sqlite

Usage examples:
  rtnpy fetch metadata
  rtnpy fetch download
  rtnpy fetch latest
  rtnpy export excel
  rtnpy export sqlite
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sqlite3
from datetime import date
from pathlib import Path
from typing import Dict

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from rtnpy import (
    Tbl,
    download_latest_file,
    extract_publication_metadata,
    fetch_publications_metadata,
    read_all_sheets,
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 "
        "Safari/537.36 "
        "Edg/131.0.0.0"
    )
}

DEFAULT_DATA_DIR = Path("data")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")

DATA_COLUMNS = ["key", "date", "value"]
HIERARCHY_COLUMNS = (
    ["key", "sheet", "sheet_title", "account_code", "account_name", "account_level"]
    + [f"P_{i}" for i in range(1, 11)]
)


def make_period_date(year, month, quarter) -> date | None:
    if year is None:
        return None
    if quarter is not None:
        return date(int(year), (int(quarter) - 1) * 3 + 1, 1)
    if month is not None:
        return date(int(year), int(month), 1)
    return date(int(year), 1, 1)


def write_header(ws, columns: list[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def get_column_names(table: Tbl) -> list[str]:
    return [col[0] for col in table.data]


def cmd_metadata(args: argparse.Namespace) -> int:
    """Fetch metadata HTML and generate metadata.json."""
    dest = Path(args.dest)
    out_html = dest / "metadata.html"
    out_json = dest / "metadata.json"

    if out_html.exists() and not args.force:
        with out_html.open("r", encoding=args.encoding) as f:
            metadata_html = f.read()
    else:
        metadata_html = fetch_publications_metadata()
        out_html.parent.mkdir(parents=True, exist_ok=True)
        with out_html.open("w", encoding=args.encoding) as f:
            f.write(metadata_html)

    soup = BeautifulSoup(metadata_html, "html.parser")
    publications = extract_publication_metadata(soup)

    out_json.parent.mkdir(parents=True, exist_ok=True)
    with out_json.open("w", encoding=args.encoding) as f:
        json.dump(publications, f, ensure_ascii=False, indent=2)

    print(f"Saved {len(publications)} publications to {out_json}")
    return 0


async def _download_link_async(
    pub: Dict,
    filename: str,
    url: str,
    client: httpx.AsyncClient,
    dest_root: Path,
    encoding: str,
) -> None:
    """Asynchronously download a single publication link to disk."""
    ano = pub.get("ano_publicacao")
    mes = pub.get("mes_publicacao")
    dest_dir = dest_root / f"{ano}-{mes:0>2}"
    dest_dir.mkdir(parents=True, exist_ok=True)

    dest_file = dest_dir / filename
    if dest_file.exists():
        return

    print(f"Downloading {url} -> {dest_file}")
    r = await client.get(url, follow_redirects=True)
    content_type = r.headers.get("Content-Type", "")
    if content_type.startswith("text/html"):
        soup = BeautifulSoup(r.text, "html.parser")
        iframe = soup.find("iframe")
        if iframe and iframe.get("src"):
            iframe_src = iframe["src"]
            print(f"Following iframe src: {iframe_src}")
            r2 = await client.get(iframe_src)
            data = r2.content
            await asyncio.to_thread(dest_file.write_bytes, data)
        else:
            await asyncio.to_thread(dest_file.write_text, r.text, encoding)
    else:
        await asyncio.to_thread(dest_file.write_bytes, r.content)


def cmd_download(args: argparse.Namespace) -> int:
    """Download files referenced in metadata.json."""
    metadata_path = Path(args.metadata)
    if not metadata_path.exists():
        print(f"Metadata file not found: {metadata_path}")
        return 2

    with metadata_path.open("r", encoding=args.encoding) as f:
        publications = json.load(f)

    async def _run_async_downloads() -> None:
        semaphore = asyncio.Semaphore(args.concurrency)

        async with httpx.AsyncClient(headers=HEADERS, timeout=600) as client:
            tasks = []

            async def _bounded(pub: Dict, filename: str, url: str) -> None:
                async with semaphore:
                    try:
                        await _download_link_async(
                            pub,
                            filename,
                            url,
                            client,
                            Path(args.dest),
                            args.encoding,
                        )
                    except Exception as exc:
                        print(f"Error downloading {url}: {exc}")

            for pub in publications:
                for filename, url in pub.get("links", {}).items():
                    tasks.append(_bounded(pub, filename, url))

            if tasks:
                await asyncio.gather(*tasks)

    asyncio.run(_run_async_downloads())
    return 0


def cmd_latest(args: argparse.Namespace) -> int:
    """Download latest single file."""
    dest = Path(args.dest)
    dest.mkdir(parents=True, exist_ok=True)
    filepath = download_latest_file(dest)
    if filepath:
        print(f"Downloaded: {filepath}")
        return 0
    else:
        print("File already exists")
        return 0


def cmd_export_excel(args: argparse.Namespace) -> int:
    """Export RTN data to Excel."""
    data_dir = Path(args.data_dir)
    data_dir.mkdir(exist_ok=True)

    print("Checking for latest RTN file...")
    filepath = download_latest_file(data_dir)

    if filepath is None:
        xlsx_files = sorted(data_dir.glob("rtn_*.xlsx"), reverse=True)
        if not xlsx_files:
            print("No RTN files found and could not download latest.")
            return 1
        filepath = xlsx_files[0]
        print(f"Using existing file: {filepath}")
    else:
        print(f"Downloaded: {filepath}")

    print(f"Reading all sheets from {filepath}...")
    results = read_all_sheets(filepath)

    output_path = Path(args.output)
    wb = Workbook()
    wb.remove(wb.active)

    ws_data = wb.create_sheet(title="rtn_data")
    write_header(ws_data, DATA_COLUMNS)
    data_row = 2

    ws_acc = wb.create_sheet(title="rtn_accounts")
    write_header(ws_acc, HIERARCHY_COLUMNS)
    acc_row = 2

    for sheet_name, (data_tbl, accounts_tbl, sheet_title) in results.items():
        print(f"Processing sheet {sheet_name}...")

        data_header = get_column_names(data_tbl)
        for row in data_tbl.iter_rows():
            if row == data_header:
                continue
            row_dict = dict(zip(data_header, row))
            account = row_dict.get("account")
            period_date = make_period_date(
                row_dict.get("year"),
                row_dict.get("month"),
                row_dict.get("quarter"),
            )
            for col_idx, value in enumerate(
                [f"{sheet_name}|{account}", period_date, row_dict.get("value")],
                start=1,
            ):
                ws_data.cell(row=data_row, column=col_idx, value=value)
            data_row += 1

        acc_header = get_column_names(accounts_tbl)
        for row in accounts_tbl.iter_rows():
            if row == acc_header:
                continue
            row_dict = dict(zip(acc_header, row))
            account_code = row_dict.get("account_code")
            p_values = [row_dict.get(f"P_{i}") for i in range(1, 11)]
            for col_idx, value in enumerate(
                [
                    f"{sheet_name}|{account_code}",
                    sheet_name,
                    sheet_title,
                    account_code,
                    row_dict.get("account_name"),
                    row_dict.get("account_level"),
                    *p_values,
                ],
                start=1,
            ):
                ws_acc.cell(row=acc_row, column=col_idx, value=value)
            acc_row += 1

    wb.save(output_path)
    print(f"Saved to {output_path}")
    return 0


def cmd_export_sqlite(args: argparse.Namespace) -> int:
    """Export RTN data to SQLite database."""
    data_dir = Path(args.data_dir)
    data_dir.mkdir(exist_ok=True)

    print("Checking for latest RTN file...")
    filepath = download_latest_file(data_dir)

    if not filepath:
        xlsx_files = sorted(data_dir.glob("rtn_*.xlsx"), reverse=True)
        if not xlsx_files:
            print("No RTN files found and could not download latest.")
            return 1
        filepath = xlsx_files[0]
        print(f"Using existing file: {filepath}")
    else:
        print(f"Downloaded: {filepath}")

    print(f"Reading all sheets from {filepath}...")
    results = read_all_sheets(filepath)

    db_path = Path(args.output)
    if db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute(
        """
        CREATE TABLE rtn_data (
            key   TEXT,
            date  TEXT,
            value REAL
        )
    """
    )

    hierarchy_cols = [f"P_{i}" for i in range(1, 11)]
    cols_sql = ", ".join(f"{col} TEXT" for col in hierarchy_cols)
    cursor.execute(
        f"""
        CREATE TABLE rtn_accounts (
            key           TEXT,
            sheet         TEXT,
            sheet_title   TEXT,
            account_code  TEXT,
            account_name  TEXT,
            account_level INTEGER,
            {cols_sql}
        )
    """
    )

    for sheet_name, (data_tbl, accounts_tbl, sheet_title) in results.items():
        print(f"Processing sheet {sheet_name}...")

        data_header = get_column_names(data_tbl)
        for row in data_tbl.iter_rows():
            if row == data_header:
                continue
            row_dict = dict(zip(data_header, row))
            account = row_dict.get("account")
            period_date = make_period_date(
                row_dict.get("year"),
                row_dict.get("month"),
                row_dict.get("quarter"),
            )
            cursor.execute(
                "INSERT INTO rtn_data (key, date, value) VALUES (?, ?, ?)",
                (f"{sheet_name}|{account}", period_date, row_dict.get("value")),
            )

        acc_header = get_column_names(accounts_tbl)
        for row in accounts_tbl.iter_rows():
            if row == acc_header:
                continue
            row_dict = dict(zip(acc_header, row))
            account_code = row_dict.get("account_code")
            p_values = [row_dict.get(f"P_{i}") for i in range(1, 11)]
            cursor.execute(
                f"INSERT INTO rtn_accounts (key, sheet, sheet_title, account_code, account_name, account_level, {', '.join(hierarchy_cols)}) VALUES (?, ?, ?, ?, ?, ?, {', '.join(['?'] * 10)})",
                [
                    f"{sheet_name}|{account_code}",
                    sheet_name,
                    sheet_title,
                    account_code,
                    row_dict.get("account_name"),
                    row_dict.get("account_level"),
                    *p_values,
                ],
            )

    conn.commit()
    conn.close()
    print(f"Database saved to {db_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="rtnpy",
        description="RTN data helper CLI - fetch and export utilities",
    )
    sub = parser.add_subparsers(dest="group", required=True)

    # Fetch subgroup
    p_fetch = sub.add_parser("fetch", help="Fetch metadata and download files")
    fetch_sub = p_fetch.add_subparsers(dest="command", required=True)

    p_meta = fetch_sub.add_parser(
        "metadata",
        help="Fetch metadata HTML and generate metadata.json",
    )
    p_meta.add_argument(
        "--dest",
        default=str(DEFAULT_DATA_DIR),
        help="Destination directory (metadata.html and metadata.json saved here)",
    )
    p_meta.add_argument("--encoding", default="utf-8", help="File encoding")
    p_meta.add_argument(
        "--force",
        action="store_true",
        help="Refetch HTML even if it exists",
    )
    p_meta.set_defaults(func=cmd_metadata)

    p_dl = fetch_sub.add_parser(
        "download",
        help="Download files referenced in metadata.json",
    )
    p_dl.add_argument(
        "--metadata",
        default=str(DEFAULT_DATA_DIR / "metadata.json"),
        help="Input metadata JSON path",
    )
    p_dl.add_argument(
        "--dest",
        default=str(DEFAULT_DATA_DIR),
        help="Destination root directory",
    )
    p_dl.add_argument(
        "--encoding",
        default="utf-8",
        help="File encoding for HTML files",
    )
    p_dl.add_argument(
        "--concurrency",
        type=int,
        default=4,
        help="Maximum concurrent downloads (default: 4)",
    )
    p_dl.set_defaults(func=cmd_download)

    p_latest = fetch_sub.add_parser("latest", help="Download latest single file")
    p_latest.add_argument(
        "--dest",
        default=str(DEFAULT_DATA_DIR),
        help="Destination directory",
    )
    p_latest.set_defaults(func=cmd_latest)

    # Export subgroup
    p_export = sub.add_parser("export", help="Export RTN data to various formats")
    export_sub = p_export.add_subparsers(dest="command", required=True)

    p_excel = export_sub.add_parser("excel", help="Export RTN data to Excel")
    p_excel.add_argument(
        "--data-dir",
        default=str(DEFAULT_DATA_DIR),
        help="Directory containing RTN files",
    )
    p_excel.add_argument(
        "--output",
        default="rtn_processed.xlsx",
        help="Output Excel file path",
    )
    p_excel.set_defaults(func=cmd_export_excel)

    p_sqlite = export_sub.add_parser("sqlite", help="Export RTN data to SQLite")
    p_sqlite.add_argument(
        "--data-dir",
        default=str(DEFAULT_DATA_DIR),
        help="Directory containing RTN files",
    )
    p_sqlite.add_argument(
        "--output",
        default="rtn_data.db",
        help="Output SQLite database path",
    )
    p_sqlite.set_defaults(func=cmd_export_sqlite)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
